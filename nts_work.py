# -------------------------------------------------------------------------------
# Name:        module1
# Purpose:
#
# Author:      Jason.Zhang
#
# Created:     30/11/2018
# Copyright:   (c) Jason.Zhang 2018
# Licence:     <your licence>
# -------------------------------------------------------------------------------
from openpyxl import load_workbook
from collections import Counter
from openpyxl.utils import get_column_letter


def main():
    pass


if __name__ == '__main__':
    main()

# change the file name to work on target file
# wb1 is for data, format, style checking
# wb2 is for formula checking
wb1 = load_workbook("C:\\Users\\Jason.Zhang\\Dropbox\\NTS-\\2019\\Q2\\First Review\\"
                   + "table_01_11PG1-3.xlsx", data_only=True)
wb2 = load_workbook("C:\\Users\\Jason.Zhang\\Dropbox\\NTS-\\2019\\Q2\\First Review\\"
                   + "table_01_11PG1-3.xlsx", data_only=False)
# since wb1 and wb2 have the same sheets and sheet names, list and names will be shared here
sheet_list = wb1.sheetnames
# sheet_list[0]: the first sheet you want to check
# sheet_list[1]: the second sheet you want to check, etc
sheet_name1 = sheet_list[0]
sheet_name2 = sheet_list[1]
# the only thing need to be separated is the sheet object
wb1_sheet1 = wb1[sheet_name1]
wb1_sheet2 = wb1[sheet_name2]
wb2_sheet1 = wb2[sheet_name1]
wb2_sheet2 = wb2[sheet_name2]

# sometimes the max_row returns a large number, temporary solution is
# to manually set the row count here, can be improved in the future
max_row1 = wb1_sheet1.max_row
max_col1 = wb1_sheet1.max_column
max_row2 = wb1_sheet2.max_row
max_col2 = wb1_sheet2.max_column
# example for manually setting the row count
# max_row1 = 58
print('\n')
print("Sheet name: " + sheet_name1 + "," + " possible incorrect row height and bold font:")
# check sheet bold font and row height
row_height_list = []
row_font_list = []
for i in range(1, max_row1+1):
    bold_list = []
    for j in range(1, max_col1+1):
        cell = wb1_sheet1.cell(row=i, column=j)
        # if the cell is not empty, then get the property and append to list
        if cell.value:
            bold_list.append(str(cell.font.bold))
    # if the list's unique value is greater than 1, which means this row has both bold and non-bold fonts
    if len(set(bold_list)) > 1:
        row_font_list.append("R" + str(i))
    # check the row height here:
    if wb1_sheet1.row_dimensions[i].height:
        if (wb1_sheet1.row_dimensions[i].height % 16.5 != 0) and (wb1_sheet1.row_dimensions[i].height % 12.75 != 0):
            row_height_list.append("R" + str(i) + "(" + str(wb1_sheet1.row_dimensions[i].height) + ")")
if row_height_list:
    print("The following rows' height is not the standard height of either Data row or Note row:")
    row_number = ""
    for row in row_height_list:
        row_number += row + " "
    print(row_number)
if row_font_list:
    print("The following rows have both bold and non-bold fonts, please check:")
    row_number = ""
    for row in row_font_list:
        row_number += row + " "
    print(row_number)
# check sheet font names
print('\n')
print("Sheet name: " + sheet_name1 + "," + " possible incorrect font row:")
for i in range(1, max_row1+1):
    font_name_list = []
    for j in range(1, max_col1+1):
        cell = wb1_sheet1.cell(row=i, column=j)
        # if the cell is not empty, then get the property and append to list
        if cell.value:
            font_name_list.append(cell.font.name)
    # if the list is not empty, then print it
    if font_name_list:
        print("Row " + str(i) + ", font name information: " + Counter(font_name_list).__str__())

# check sheet font sizes
print('\n')
print("Sheet name: " + sheet_name1)
for i in range(1, max_row1+1):
    font_size_list = []
    for j in range(1, max_col1+1):
        cell = wb1_sheet1.cell(row=i, column=j)
        # if the cell is not empty, then get the property and append to list
        if cell.value:
            font_size_list.append(int(cell.font.size))
    # if the list is not empty, then print it
    if font_size_list:
        print("Row " + str(i) + ", font size information: " + Counter(font_size_list).__str__())

# check sheet highlight cells
print('\n')
print("Sheet name: " + sheet_name2)
print("You may need to double check the following highlighted cells: ")
for i in range(1, max_row2 + 1):
    incorrect_highlight_list = []
    # don't check the last column because it should always be highlighted(supposed to be new column)
    for j in range(1, max_col2):
        cell = wb1_sheet2.cell(row=i, column=j)
        # FFFFFF00--yellow highlighted, 00000000--no highlight
        if (cell.fill.start_color.rgb != "00000000" and str(cell.value) == "TRUE") or \
                (cell.fill.start_color.rgb == "00000000" and str(cell.value) == "FALSE"):
            incorrect_highlight_cell = "[" + str(i) + "|" + get_column_letter(j) + "]"
            incorrect_highlight_list.append(incorrect_highlight_cell)
        # if the list is not empty, then print it
    if incorrect_highlight_list:
        # the first column should be excluded
        if len(incorrect_highlight_list) == max_col2 - 1:
            print("All cells in row" + " " + str(i))
        else:
            cell_index = ""
            for item in incorrect_highlight_list:
                cell_index += item + " "
            print(cell_index)

# formula checking starts here
print('\n')
print("Sheet name: " + sheet_name1)
print("The following cell(s) has formula(s) on it, please check the previous' version to see if they match:")
for i in range(1, max_row1+1):
    formula_list = []
    for j in range(1, max_col1+1):
        cell = wb2_sheet1.cell(row=i, column=j)
        # if the cell has formula on it, then record it since sheet1 is not supposed to have formulas
        if cell.data_type == "f":
            incorrect_formula_cell = "[" + str(i) + "|" + get_column_letter(j) + "]"
            formula_list.append(incorrect_formula_cell)
    # if the list is not empty, then print it
    if formula_list:
        # the first column should be excluded
        if len(formula_list) == max_col1 - 1:
            print("All cells in row" + " " + str(i))
        else:
            cell_index = ""
            for item in formula_list:
                cell_index += item + " "
            print(cell_index)
# formula pattern check for working sheet
# TODO, this needs to think over and develop a good algorithm
# print('\n')
# print("Sheet name: " + sheet_name2)
# print("Formula patterns in working sheet: ")
# for i in range(1, max_row1+1):
#     formula_list = []
#     for j in range(1, max_col1+1):
#         cell = wb2_sheet2.cell(row=i, column=j)
#         # if the cell has formula on it, then record it since sheet1 is not supposed to have formulas
#         if cell.data_type == "f":
#             formula_list.append(cell.value)
#     # if the list is not empty, then print it
#     if formula_list:
#         formula = ""
#         for item in formula_list:
#             formula += item + " "
#         print(formula)



